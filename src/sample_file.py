# sample_file.py    12May2020  crs
""" Gather samples from custome .xlsx file
"""
import sys
import openpyxl
import re 

from select_trace import SlTrace

class SamplePoint:
    def __init__(self, lat=None, long=None, plot_key=None):
        self.lat = lat
        self.long = long
        self.plot_key = plot_key
        
    def __str__(self):
        return f"SamplePoint: {self.plot_key} lat:{self.lat} Long:{self.long}"

    def get_plot_key(self):
        return self.plot_key

    def latLong(self):
        return self.lat, self.long
        
class SampleFile:
    """ simple point loading 
    """
    def __init__(self, file_name=None):
        self.file_name = file_name
        self.points = []
        if file_name is not None:
            self.load_file(file_name)
    
    def load_file(self, file_name):
        """ load .xlsx file with sample points
        :file_name: file name
        """
        if file_name is None:
            file_name = self.file_name  # Use stored
        self.file_name = file_name      # Save name
        wb = openpyxl.load_workbook(file_name)
        sheet = wb['WhitneyHill_average']
        nrow = sheet.max_row
        SlTrace.lg("%d rows" % nrow)
        ncol = sheet.max_column
        SlTrace.lg("%d columns" % ncol)
        point_header = "POINT"
        plot_header = "Plot"
        long_header = "long_deg"
        lat_header = "lat_deg"
        max_long = None
        min_long = None
        max_lat = None
        min_lat = None
        """
        Find column headers and beginning of data rows
        """
        
        got_headers = False
        header_row = None
        point_colno = None
        plot_colno = None
        long_colno = None
        lat_colno = None
        for nr in range(1, nrow+1):
            value = sheet.cell(row=nr, column=1).value
            if value != "POINT":
                continue            # Look at next row
            
            for nc in range(1, ncol+1):
                value = sheet.cell(row=nr, column=nc).value
                if value == point_header:
                    point_colno = nc
                elif value == plot_header:
                    plot_colno = nc
                elif value == long_header:      # Finds last one in row
                    long_colno = nc
                elif value == lat_header:
                    lat_colno = nc
                    
            if point_colno is None:
                SlTrace.lg("POINT column missing")
                sys.exit(1)
            if plot_colno is None:
                SlTrace.lg("Plot column missing")
                sys.exit(1)
            if long_colno is None:
                SlTrace.lg("long column missing")
                sys.exit(1)
            if lat_colno is None:
                SlTrace.lg("lat column missing")
                sys.exit(1)
            header_row = nr
            break
        
        if header_row is None:
            SlTrace.lg("Header row not found")
            sys.exit(1)
        
        """
        Collect points and find extent of lat,long
        """
        points = []
        limit_pointh = {}   # Hash by limit
        pointh = {}         # Hash by plot
        lat_colno_1 = lat_colno - 3     # Orig #'s are 3 cols left
        long_colno_1 = long_colno - 3
        for nr in range(header_row+1, nrow+1):
            lat = sheet.cell(row=nr, column=lat_colno).value
            if lat is None:
                lat = sheet.cell(row=nr, column=lat_colno_1).value
                
            long = sheet.cell(row=nr, column=long_colno).value
            if long is None:
                long = sheet.cell(row=nr, column=long_colno_1).value
                
            if lat is None or long is None:
                continue
            
            plot = sheet.cell(row=nr, column=plot_colno).value
            pm = re.match("T(\d+)P(\d+)", plot)
            if pm is not None:
                plot_key = f"{pm.group(1)}-{pm.group(2)}"
            else:
                plot_key = plot
            if max_lat is None or lat > max_lat:
                max_lat = lat
                limit_pointh['max_lat'] = plot
            if min_lat is None or lat < min_lat:
                min_lat = lat
                limit_pointh['min_lat'] = plot
            if max_long is None or long > max_long:
                max_long = long
                limit_pointh['max_long'] = plot
            if min_long is None or long < min_long:
                min_long = long
                limit_pointh['min_long'] = plot
            point = SamplePoint(plot_key=plot_key, lat=float(lat), long=float(long))
            self.points.append(point)
            pointh[plot] = point
            
        SlTrace.lg("%d Sample Points" % len(points))
        SlTrace.lg("Max Longitude: %.5f Latitude: %.5f" % (max_long, max_lat))
        SlTrace.lg("Min Longitude: %.5f Latitude: %.5f" % (min_long, min_lat))
        SlTrace.lg("Points on the edge")
        for key in limit_pointh.keys():
            plot = limit_pointh[key]
            point = pointh[plot]
            pm = re.match("T(\d+)P(\d+)", plot)
            if pm is not None:
                plot_key = f"{pm.group(1)}-{pm.group(2)}"
            else:
                plot_key = plot
            SlTrace.lg("%-8s %s  Longitude: %.5f latitude: %.5f" % (key, plot_key, point.long, point.lat))
    
    def get_point(self, plot_key):
        """ Get point, given plot_key
        :plot_key: point plot key
        :returns: point if key plot_key found
                else None
        """
        for point in self.points:
            pk = point.get_plot_key()
            if pk == plot_key:
                return point
            
        return None
    
    def get_points(self):
        """ Return list of SamplePoint points
        """
        return self.points
