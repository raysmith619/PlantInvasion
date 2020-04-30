#! sample_points.py
"""
Read sample points and find encompasing rectangle
"""
import sys
import os
import re
import datetime
import openpyxl
import argparse
from PIL import Image, ImageDraw, ImageFont

from select_trace import SlTrace
from GoogleMapImage import GoogleMapImage, geoDistance
from scrolled_canvas import ScrolledCanvas 

adjWidthForSize = True       # widen lines to aid visibility on large images
mapRotate = 45.
showSampleLL = True
forceNew = False
maptype = "hybrid"
scale = 1
testMarks = False
useOldFile = False
zoom = 19
parser = argparse.ArgumentParser()
parser.add_argument('--adjwidthforsize', dest='adjWidthForSize', default=adjWidthForSize, action='store_true')
parser.add_argument('--maprotate=', type=float, dest='mapRotate', default=mapRotate)
parser.add_argument('--maptype', dest='maptype', default=maptype)
parser.add_argument('--showsamplell', dest='showSampleLL', default=showSampleLL, action='store_true')
parser.add_argument('--forcenew', dest='forceNew', default=forceNew, action='store_true')
parser.add_argument('--scale=', type=int, dest='scale', default=scale)
parser.add_argument('--testmarks', dest='testMarks', default=testMarks, action='store_true')
parser.add_argument('--notestmarks', dest='testMarks', action='store_false')
parser.add_argument('--useoldfile', dest='useOldFile', default=useOldFile, action='store_true')
parser.add_argument('--zoom=', type=int, dest='zoom', default=zoom)

args = parser.parse_args()             # or die "Illegal options"
mapRotate = args.mapRotate
maptype = args.maptype
showSampleLL = args.showSampleLL
forceNew = args.forceNew
scale = args.scale
testMarks = args.testMarks
useOldFile = args.useOldFile
zoom = args.zoom
SlTrace.lg("%s %s\n" % (os.path.basename(sys.argv[0]), " ".join(sys.argv[1:])))
SlTrace.lg("args: %s\n" % args)


sample_file = "../data/2018 05 12 Revised GPS coordinates of 32 sample plot centers.xlsx"
wb = openpyxl.load_workbook(sample_file)
sheet = wb['WhitneyHill_average']
nrow = sheet.max_row
SlTrace.lg("%d rows" % nrow)
ncol = sheet.max_column
SlTrace.lg("%d columns" % ncol)
point_header = "POINT"
plot_header = "Plot"
long_header = "long_deg"
lat_header = "lat_deg"
max_long = None
min_long = None
max_lat = None
min_lat = None
"""
Find column headers and beginning of data rows
"""

got_headers = False
header_row = None
point_colno = None
plot_colno = None
long_colno = None
lat_colno = None
for nr in range(1, nrow+1):
    value = sheet.cell(row=nr, column=1).value
    if value != "POINT":
        continue            # Look at next row
    
    for nc in range(1, ncol+1):
        value = sheet.cell(row=nr, column=nc).value
        if value == point_header:
            point_colno = nc
        elif value == plot_header:
            plot_colno = nc
        elif value == long_header:      # Finds last one in row
            long_colno = nc
        elif value == lat_header:
            lat_colno = nc
            
    if point_colno is None:
        SlTrace.lg("POINT column missing")
        sys.exit(1)
    if plot_colno is None:
        SlTrace.lg("Plot column missing")
        sys.exit(1)
    if long_colno is None:
        SlTrace.lg("long column missing")
        sys.exit(1)
    if lat_colno is None:
        SlTrace.lg("lat column missing")
        sys.exit(1)
    header_row = nr
    break

if header_row is None:
    SlTrace.lg("Header row not found")
    sys.exit(1)

"""
Collect points and find extent of lat,long
"""
points = []
limit_pointh = {}   # Hash by limit
pointh = {}         # Hash by plot
lat_colno_1 = lat_colno - 3     # Orig #'s are 3 cols left
long_colno_1 = long_colno - 3
for nr in range(header_row+1, nrow+1):
    lat = sheet.cell(row=nr, column=lat_colno).value
    if lat is None:
        lat = sheet.cell(row=nr, column=lat_colno_1).value
        
    long = sheet.cell(row=nr, column=long_colno).value
    if long is None:
        long = sheet.cell(row=nr, column=long_colno_1).value
        
    if lat is None or long is None:
        continue
    
    plot = sheet.cell(row=nr, column=plot_colno).value
    if max_lat is None or lat > max_lat:
        max_lat = lat
        limit_pointh['max_lat'] = plot
    if min_lat is None or lat < min_lat:
        min_lat = lat
        limit_pointh['min_lat'] = plot
    if max_long is None or long > max_long:
        max_long = long
        limit_pointh['max_long'] = plot
    if min_long is None or long < min_long:
        min_long = long
        limit_pointh['min_long'] = plot
    point = {'plot' : plot, 'lat' : lat, 'long' : long}
    points.append(point)
    pointh[plot] = point
    
SlTrace.lg("%d Sample Points" % len(points))
SlTrace.lg("Max Longitude: %.5f Latitude: %.5f" % (max_long, max_lat))
SlTrace.lg("Min Longitude: %.5f Latitude: %.5f" % (min_long, min_lat))
SlTrace.lg("Points on the edge")
for key in limit_pointh.keys():
    plot = limit_pointh[key]
    point = pointh[plot]
    pm = re.match("T(\d+)P(\d+)", plot)
    if pm is not None:
        plot_key = f"{pm.group(1)}-{pm.group(2)}"
    else:
        plot_key = plot
    SlTrace.lg("%-8s %s  Longitude: %.5f latitude: %.5f" % (key, plot_key, point['long'], point['lat']))

"""
Plot using points, mapRotate, and mapBorder as guide
"""

gmi = GoogleMapImage( mapPoints=points,
                      mapBorderM=400,
                      showSampleLL=showSampleLL,
                      scale=scale,
                      forceNew=forceNew,
                      maptype=maptype, zoom=zoom,
                      mapRotate=mapRotate,
                      useOldFile=useOldFile)
SlTrace.lg("Plot boundaries")
SlTrace.lg("    ulLat=%.5f ulLong=%.5f lrLat=%.5f lrLong=%.5f" %
                    (gmi.ulLat, gmi.ulLong, gmi.lrLat, gmi.lrLong))
SlTrace.lg("    corner to corner:= %.2f meters" %
                     geoDistance(latLong=(gmi.ulLat, gmi.ulLong), latLong2=(gmi.lrLat, gmi.lrLong)))
SlTrace.lg("    image %d width in pixels %d height in pixels" % (gmi.image.width, gmi.image.height))
    


"""
Add samples to plot
"""
test_point = 0
for point in points:
    test_point += 1
    gmi.addSample(point)
    if test_point >= 4:
        pass


if testMarks:
    """   
    A center point surrounded with N,E,S, and W markers
    """
    blue = (0,0,255)
    red = (255, 0, 0)
    t_size = 20
    t_color = red
    t_font = ImageFont.truetype("arial.ttf", size=t_size)
    lat_len = abs((gmi.ulLat-gmi.lrLat)/8)
    long_len = abs((gmi.lrLong-gmi.ulLong)/8)
    SlTrace.lg("lat_len=%.6f long_len=%.6f" % (lat_len, long_len))
    c_latLong = ((gmi.ulLat+gmi.lrLat)/2, (gmi.ulLong+gmi.lrLong)/2)
    n_latLong = (c_latLong[0]+lat_len, c_latLong[1])
    e_latLong = (c_latLong[0], c_latLong[1]+long_len)
    s_latLong = (c_latLong[0]-lat_len, c_latLong[1])
    w_latLong = (c_latLong[0], c_latLong[1]-long_len)
    SlTrace.lg("North %.6f, %.6f" % (n_latLong[0], n_latLong[1]))
    SlTrace.lg("East %.6f, %.6f" % (e_latLong[0], e_latLong[1]))
    SlTrace.lg("South %.6f, %.6f" % (s_latLong[0], s_latLong[1]))
    SlTrace.lg("West %.6f, %.6f" % (w_latLong[0], w_latLong[1]))
    SlTrace.lg("Center %.6f, %.6f" % (c_latLong[0], c_latLong[1]))
    gmi.text("North %.6f, %.6f" % (n_latLong[0], n_latLong[1]), latLong=n_latLong, font=t_font, fill=t_color)
    gmi.text("East %.6f, %.6f" % (e_latLong[0], e_latLong[1]), latLong=e_latLong, font=t_font, fill=t_color)
    gmi.text("South %.6f, %.6f" % (s_latLong[0], s_latLong[1]), latLong=s_latLong, font=t_font, fill=t_color)
    gmi.text("West %.6f, %.6f" % (w_latLong[0], w_latLong[1]), latLong=w_latLong, font=t_font, fill=t_color)
    gmi.text("Center %.6f, %.6f" % (c_latLong[0], c_latLong[1]), latLong=c_latLong, font=t_font, fill=t_color)
    """
    Do box around orignial boxex
    """
    gd = gmi.geoDraw
    inset = 100
    box_line_width = 10
    gd.line([gd.addToPoint(latLong=(max_lat, min_long), leng=inset, deg=-45),
              gd.addToPoint(latLong=(max_lat, max_long), leng=inset, deg=-180+45),
            gd.addToPoint(latLong=(min_lat, max_long), leng=inset, deg=-180-45),
             gd.addToPoint(latLong=(min_lat, min_long), leng=inset, deg=45),
            gd.addToPoint(latLong=(max_lat, min_long), leng=inset, deg=-45)],
              fill=blue, width=box_line_width)

        
        
        ###break       # limit for debugging TFD
if testMarks:
    ###gmi.addScale(latLong=(42.37350, -71.18318), len=100, marks=10, unitName="f")
    ###gmi.addScale(latLong=(42.37365, -71.18304), deg=8, len=100, marks=10, unitName="f", color=(0,0,255))
    ###gmi.addScale(latLong=(42.37365, -71.183060), deg=8.8, len=500, marks=10, unitName="f", color=(0,0,255))
    gmi.addScale(latLong=(42.37357, -71.183000), deg=8.8, len=500, marks=10, unitName="f", color=(0,0,255))
    gmi.addScale(latLong=(42.37555,-71.18677), latLongEnd=(42.37077,-71.18208)) # Diagonal ul to lr
    gmi.addScale(latLong=(42.37555,-71.18677), latLongEnd=(42.37077,-71.18677)) # Vertical left edge
    gmi.addScale(latLong=(42.37077,-71.18677), latLongEnd=(42.37077,-71.18208))
    gmi.addScale(xY=(10,10), deg=-90, len=gmi.getHeight()*.9)
    
gmi.addScale(deg=-mapRotate)            # Default scale in meters at bottom
gmi.addScale(xY = (gmi.getWidth()*.1, gmi.getHeight()*.95),
            leng = gmi.getWidth()*.8, unitName="f", bigMarks=10,
            deg=-mapRotate)
show_yards = False
if show_yards:
    gmi.addScale(xY = (gmi.getWidth()*.1, gmi.getHeight()*.975),
                leng = gmi.getWidth()*.8, unitName="y", bigMarks=10,
                deg=-mapRotate)
if mapRotate is not None:
    SlTrace.lg("Rotate map %.2f degrees" % mapRotate)

if mapRotate is None:
    rotate = 0.
else:
    rotate = mapRotate
now = datetime.datetime.now().strftime("%b %d %Y %H:%M:%S")

title = "rotate = %.2f\n%s" % (rotate, now)
gmi.addTitle(title)

tennis_court = True
if tennis_court:
    gd = gmi.geoDraw
    ptc = gmi.getXY(latLong=(42.37332, -71.18349))
    th_deg = -mapRotate+8.0
    nw_corner_1 = gd.addToPoint(xY=ptc, leng=gd.meterToPixel(110), deg=th_deg)
    nw_corner = gd.addToPoint(xY=nw_corner_1, leng=gd.meterToPixel(54), deg=th_deg+90)
    ###pthird = gd.addToPoint(xY=pthird, leng=gd.meterToPixel(0.182), deg=th_deg)
    gmi.addScale(xY=nw_corner, deg=th_deg, unitName='f', leng=gd.meterToPixel(100), tic_dir=-1)


third_to_home = True
if third_to_home:
    gd = gmi.geoDraw
    p3_1 = gmi.getXY(latLong=(42.37332, -71.18349))
    th_deg = -mapRotate+9.3
    pthird = gd.addToPoint(xY=p3_1, leng=gd.meterToPixel(45.85), deg=90-6.4-mapRotate)
    pthird = gd.addToPoint(xY=pthird, leng=gd.meterToPixel(0.182), deg=th_deg)
    gmi.addScale(xY=pthird, deg=th_deg, unitName='f', leng=gd.meterToPixel(50), tic_dir=-1)

football_sideline = True
if football_sideline:
    gd = gmi.geoDraw
    goal_p = gmi.getXY(latLong=(42.3736665, -71.18350))
    th_deg = -mapRotate+9.5+90
    goal_left_p = gd.addToPoint(xY=goal_p, leng=gd.meterToPixel(56.5), deg=th_deg)
    goal_left_p = gd.addToPoint(xY=goal_left_p, leng=gd.meterToPixel(-6.6), deg=th_deg+90)
    gmi.addScale(xY=goal_left_p, deg=th_deg, unitName='y', leng=gd.meterToPixel(100))

###gmi.show()
aug_name = gmi.saveAugmented()
sc = ScrolledCanvas(fileName=aug_name)
sc.mainloop()

        